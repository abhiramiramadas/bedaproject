# ==========================================================
# Excel Export for Accident Data
# ==========================================================

import os
from datetime import datetime
from typing import List, Dict, Optional, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    Workbook = None
    Font = Alignment = PatternFill = Border = Side = None
    get_column_letter = None


class ExcelExporter:
    """Export accident data to Excel files"""
    
    def __init__(self, output_dir: str = "uploads/exports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        
        if EXCEL_AVAILABLE:
            # Style definitions
            self.header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
            self.critical_fill = PatternFill(start_color="dc3545", end_color="dc3545", fill_type="solid")
            self.high_fill = PatternFill(start_color="ff6600", end_color="ff6600", fill_type="solid")
            self.medium_fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
            self.low_fill = PatternFill(start_color="00d26a", end_color="00d26a", fill_type="solid")
            
            self.header_font = Font(bold=True, color="ffffff", size=11)
            self.title_font = Font(bold=True, size=16, color="1a1a2e")
            
            self.thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    def export_accidents(self, accidents: List[Dict], filename: Optional[str] = None) -> Optional[str]:
        """
        Export accidents to Excel file
        
        Args:
            accidents: List of accident dictionaries
            filename: Optional filename (auto-generated if not provided)
            
        Returns:
            Path to exported file or None if unavailable
        """
        if not EXCEL_AVAILABLE:
            print("Excel export not available - openpyxl not installed")
            return None
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"Accident_Report_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Accidents"
        
        # Title
        ws.merge_cells('A1:L1')
        ws['A1'] = "AI Accident Detection System - Accident Report"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Subtitle with date
        ws.merge_cells('A2:L2')
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = [
            "ID", "Timestamp", "Severity", "Impact Score", "IoU %",
            "Vehicles", "Latitude", "Longitude", "Address",
            "Hospital", "Police", "Weather"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.thin_border
        
        # Data rows
        for row_num, accident in enumerate(accidents, 5):
            # ID
            ws.cell(row=row_num, column=1, value=accident.get('id', row_num - 4))
            
            # Timestamp
            timestamp_str = accident.get('timestamp', '')
            if timestamp_str:
                try:
                    dt = datetime.fromisoformat(timestamp_str.replace('Z', '+00:00'))
                    timestamp_str = dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    pass
            ws.cell(row=row_num, column=2, value=timestamp_str)
            
            # Severity with color
            severity = accident.get('severity', 'UNKNOWN')
            cell = ws.cell(row=row_num, column=3, value=severity)
            if severity == 'CRITICAL':
                cell.fill = self.critical_fill
                cell.font = Font(bold=True, color="ffffff")
            elif severity == 'HIGH':
                cell.fill = self.high_fill
                cell.font = Font(bold=True, color="ffffff")
            elif severity == 'MEDIUM':
                cell.fill = self.medium_fill
                cell.font = Font(bold=True)
            elif severity == 'LOW':
                cell.fill = self.low_fill
                cell.font = Font(bold=True)
            
            # Impact Score
            ws.cell(row=row_num, column=4, value=round(accident.get('impact_score', 0), 1))
            
            # IoU
            iou = accident.get('iou', 0)
            ws.cell(row=row_num, column=5, value=f"{iou:.1%}" if iou else "N/A")
            
            # Vehicles
            vehicles = accident.get('vehicle_types', [])
            if isinstance(vehicles, list):
                vehicles = ', '.join(vehicles)
            ws.cell(row=row_num, column=6, value=vehicles)
            
            # Location
            ws.cell(row=row_num, column=7, value=accident.get('latitude', 0))
            ws.cell(row=row_num, column=8, value=accident.get('longitude', 0))
            ws.cell(row=row_num, column=9, value=str(accident.get('address', 'N/A'))[:50])
            
            # Emergency services
            ws.cell(row=row_num, column=10, value=accident.get('nearest_hospital', 'N/A'))
            ws.cell(row=row_num, column=11, value=accident.get('nearest_police', 'N/A'))
            
            # Weather
            ws.cell(row=row_num, column=12, value=accident.get('weather', 'N/A'))
            
            # Apply borders
            for col in range(1, 13):
                ws.cell(row=row_num, column=col).border = self.thin_border
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        column_widths = [6, 20, 12, 12, 10, 25, 12, 12, 40, 25, 25, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Add summary sheet
        self._add_summary_sheet(wb, accidents)
        
        # Save
        wb.save(filepath)
        print(f"Excel report exported: {filepath}")
        
        return filepath
    
    def _add_summary_sheet(self, wb: Any, accidents: List[Dict]) -> None:
        """Add summary statistics sheet"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = "Accident Summary Statistics"
        ws['A1'].font = self.title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Stats
        total = len(accidents)
        critical = sum(1 for a in accidents if a.get('severity') == 'CRITICAL')
        high = sum(1 for a in accidents if a.get('severity') == 'HIGH')
        medium = sum(1 for a in accidents if a.get('severity') == 'MEDIUM')
        low = sum(1 for a in accidents if a.get('severity') == 'LOW')
        
        stats = [
            ("Total Accidents", total),
            ("Critical", critical),
            ("High", high),
            ("Medium", medium),
            ("Low", low),
            ("", ""),
            ("Average Impact Score", round(sum(a.get('impact_score', 0) for a in accidents) / max(total, 1), 1)),
        ]
        
        for row, (label, value) in enumerate(stats, 3):
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15


# Global instance - only create if available
excel_exporter = ExcelExporter() if EXCEL_AVAILABLE else None


def export_to_excel(accidents: List[Dict], filename: Optional[str] = None) -> Optional[str]:
    """Export accidents to Excel file"""
    if excel_exporter:
        return excel_exporter.export_accidents(accidents, filename)
    return None


def export_accidents_to_excel(accidents: List[Dict], filename: Optional[str] = None) -> Optional[str]:
    """Export accidents to Excel file (alias)"""
    return export_to_excel(accidents, filename)
